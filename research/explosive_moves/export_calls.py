"""Export the full trade log of Risk-Adjusted Momentum (5cr+, top-25, monthly) from 2019.

Writes CSVs (always) + an .xlsx if openpyxl is present:
  - All Calls       : every position (stock, buy, sell, months, return)
  - Monthly Holdings: the 25 names held each month, with a NEW-ENTRY flag + that month's return
  - Monthly Turnover: positions / buys / sells / portfolio return each month
  - Summary         : headline stats + turnover
Also prints the turnover numbers (trades per month).
"""
import csv
import numpy as np
from pathlib import Path
from explosive_moves.embase import load_symbol_cache
from explosive_moves.strategies import CR
from explosive_moves.metrics import index_series
from explosive_moves.factory import build_tables, sig_riskadj, eqstats, COST_PS

OUT = Path("/opt/hermes/research/explosive_moves/out")
OUT.mkdir(exist_ok=True)
TOPN = 25
REBAL = 22

cache = load_symbol_cache()
dts, cl = index_series("Nifty 50")
cal = [d for d in dts if d >= "2012-06-01"]
rebal_idx = list(range(0, len(cal), REBAL))
tables = build_tables(cache, cal, rebal_idx, 5 * CR)

held = {}            # sym -> [entry_date, factor, months]
closed = []          # (sym, entry, exit, months, ret)
holdings = []        # (month, stock, is_new, month_ret)
turnover = []        # (month, n_held, n_buys, n_sells, port_ret)
mrets = []
prev = set()
for t in tables:
    if t["d0"] < "2019-01-01":
        continue
    m = t["d0"]
    score, _ = sig_riskadj(t)
    valid = ~np.isnan(score) & ~np.isnan(t["fwd"])
    idx = np.where(valid)[0]
    order = idx[np.argsort(score[idx])[::-1]][:TOPN]
    picks = {t["syms"][i]: float(t["fwd"][i]) for i in order}
    pset = set(picks)
    buys, sells = pset - prev, prev - pset
    # close dropped
    for sym in list(held):
        if sym not in picks:
            closed.append((sym, held[sym][0], m, held[sym][2], held[sym][1] * (1 - COST_PS) - 1))
            del held[sym]
    # open new
    for sym in picks:
        if sym not in held:
            held[sym] = [m, (1 - COST_PS), 0]
    # accrue + record holdings
    for sym, fwd in picks.items():
        held[sym][1] *= (1 + fwd)
        held[sym][2] += 1
        holdings.append((m, sym, 1 if sym in buys else 0, round(fwd * 100, 1)))
    g = float(np.mean(list(picks.values())))
    turn = (len(buys) + len(sells)) / max(len(pset), 1)
    pr = g - COST_PS * ((len(buys) + len(sells)) / max(len(pset), 1))
    mrets.append(pr)
    turnover.append((m, len(pset), len(buys), len(sells), round(pr * 100, 2)))
    prev = pset
for sym, (en, fac, mo) in held.items():
    closed.append((sym, en, cal[rebal_idx[-1]], mo, fac * (1 - COST_PS) - 1))

closed.sort(key=lambda c: -c[4])
s = eqstats(np.array(mrets))
nbuys = np.mean([r[2] for r in turnover]); nsells = np.mean([r[3] for r in turnover])
nwin = sum(1 for c in closed if c[4] > 0)

print(f"Positions held at a time: {TOPN}")
print(f"Avg per month: {nbuys:.1f} new buys + {nsells:.1f} sells = {nbuys+nsells:.1f} trades/month")
print(f"Total calls 2019-2026: {len(closed)} | winners {nwin} ({nwin/len(closed)*100:.0f}%)")
print(f"CAGR {s['cagr']*100:.1f}%  MaxDD {s['maxdd']*100:.1f}%  total {s['totx']:.1f}x")

# --- CSVs ---
with open(OUT / "riskadj_calls.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(["stock", "buy_date", "sell_date", "months_held", "return_pct"])
    for sym, en, ex, mo, r in closed:
        w.writerow([sym, en, ex, mo, round(r * 100, 1)])
with open(OUT / "riskadj_holdings.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(["month", "stock", "new_entry", "month_return_pct"])
    for row in holdings:
        w.writerow(row)
with open(OUT / "riskadj_turnover.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(["month", "positions", "new_buys", "sells", "portfolio_return_pct"])
    for row in turnover:
        w.writerow(row)
print("CSVs written to out/.")

# --- xlsx (if openpyxl present) ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="305496")

    def sheet(ws, headers, rows):
        ws.append(headers)
        for c in ws[1]:
            c.font = hdr; c.fill = fill; c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(list(r))
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(11, min(22, max(len(str(c.value or "")) for c in col) + 2))

    ws = wb.active; ws.title = "Summary"
    sheet(ws, ["Metric", "Value"], [
        ["Strategy", "Risk-Adjusted Momentum (6m return / volatility)"],
        ["Universe", "NSE cash, >=Rs 5cr avg daily turnover"],
        ["Positions held at a time", TOPN],
        ["Rebalance", "Monthly"],
        ["Avg new buys / month", round(nbuys, 1)],
        ["Avg sells / month", round(nsells, 1)],
        ["Avg trades / month", round(nbuys + nsells, 1)],
        ["Period", "2019-01 to 2026-06"],
        ["CAGR", f"{s['cagr']*100:.1f}%"],
        ["Max drawdown", f"{s['maxdd']*100:.1f}%"],
        ["Total growth", f"{s['totx']:.1f}x"],
        ["Total calls", len(closed)],
        ["Win rate", f"{nwin/len(closed)*100:.0f}%"],
    ])
    sheet(wb.create_sheet("All Calls"), ["Stock", "Buy date", "Sell date", "Months held", "Return %"],
          [[sym, en, ex, mo, round(r * 100, 1)] for sym, en, ex, mo, r in closed])
    sheet(wb.create_sheet("Monthly Holdings"), ["Month", "Stock", "New entry (1=buy)", "Month return %"], holdings)
    sheet(wb.create_sheet("Monthly Turnover"),
          ["Month", "Positions", "New buys", "Sells", "Portfolio return %"], turnover)
    wb.save(OUT / "RiskAdjMomentum_trades_2019.xlsx")
    print("XLSX written: out/RiskAdjMomentum_trades_2019.xlsx")
except ImportError:
    print("NO_OPENPYXL")

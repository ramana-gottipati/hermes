# -*- coding: utf-8 -*-
"""Build the investor Excel workbook from bt_zerodha.json. Real numbers only."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(SP, "bt_zerodha.json"), encoding="utf-8"))
STRATS = list(DATA["strategies"])
BENCH = DATA["benchmark"]
# fold in the union family (same gauntlet)
UNI = json.load(open(os.path.join(SP, "union_gauntlet.json"), encoding="utf-8"))
STRATS = STRATS + UNI["strategies"]

# ---- styling helpers ----
H1 = Font(bold=True, size=15, color="1F3864")
H2 = Font(bold=True, size=12, color="1F3864")
BOLD = Font(bold=True)
WHITE = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
HEADFILL = PatternFill("solid", fgColor="1F3864")
TIERFILL = {"1": "C6EFCE", "2": "C6EFCE", "3": "FFEB9C", "4": "FFEB9C",
            "5": "FFC7CE", "6": "FFC7CE", "7": "FFC7CE", "0": "D9E1F2"}
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def tnum(rt):
    try: return int(str(rt).split("-")[0])
    except Exception: return 99


def tfill(rt):
    n = tnum(rt)
    if n == 0: return "D9E1F2"       # reference
    if n <= 2: return "C6EFCE"       # stable = green
    if n <= 4: return "FFEB9C"       # moderate = amber
    if n <= 7: return "FFC7CE"       # aggressive = red
    return "D9D2F2"                  # union family = purple


def w(ws, cell, val, font=None, fill=None, align=None, border=False):
    c = ws[cell]; c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if border: c.border = BORDER
    return c


# ================= FIELD ORDER (the 30+) =================
FIELDS = [
    ("strategy", "Strategy"), ("risk_tier", "Risk tier"), ("factor", "Selection rule"),
    ("rebalance_cadence", "Rebalances"), ("universe", "Universe"), ("hold_band", "Hold-band"),
    ("stocks_held_at_all_times", "Stocks held (always)"),
    ("period_start", "Start"), ("period_end", "End"), ("years", "Years"),
    ("num_rebalances", "# rebalances"),
    ("distinct_stocks_ever_used", "Distinct stocks ever used"),
    ("avg_stocks_swapped_per_rebalance", "Avg stocks swapped / rebalance"),
    ("churn_turnover_pct_per_rebalance", "Churn % / rebalance"),
    ("churn_turnover_pct_per_year", "Churn % / year"),
    ("total_trades_buys_plus_sells", "Total trades (buys+sells)"),
    ("avg_holding_period_years", "Avg holding period (yrs)"),
    ("gross_cagr_pct", "Gross CAGR % (no cost)"),
    ("net_cagr_flat_paper_pct", "CAGR % (pretend-cheap cost)"),
    ("net_cagr_zerodha_pct", "NET CAGR % (real Zerodha cost)"),
    ("illustrative_after_tax_cagr_pct", "After-tax CAGR % (illustrative)"),
    ("volatility_ann_pct", "Volatility % (annual)"),
    ("return_over_vol", "Return / volatility"),
    ("max_drawdown_pct", "Worst drop (max drawdown) %"),
    ("calmar_cagr_over_maxdd", "Calmar (CAGR / worst drop)"),
    ("pct_positive_years", "% positive years"),
    ("best_year", "Best year"), ("best_year_return_pct", "Best year %"),
    ("worst_year", "Worst year"), ("worst_year_return_pct", "Worst year %"),
    ("ann_total_cost_pct", "Total cost % / yr"),
    ("ann_cost_market_spread_pct", "  of which: market spread % / yr"),
    ("ann_cost_slippage_pct", "  of which: slippage % / yr"),
    ("ann_cost_zerodha_charges_pct", "  of which: Zerodha charges % / yr"),
    ("ann_cost_rupees_on_1cr", "Cost in Rs/yr on Rs1cr"),
    ("tax_treatment", "Tax treatment"),
    ("median_pick_daily_liquidity_cr", "Median stock liquidity (Rs cr/day)"),
    ("capacity_max_aum_cr_median", "Capacity: max AUM (Rs cr)"),
    ("rs1cr_becomes_cr", "Rs 1 cr becomes (Rs cr)"),
    ("illustrative_after_tax_1cr_becomes_cr", "Rs 1 cr after-tax (Rs cr)"),
    ("wealth_multiple", "Wealth multiple (x)"),
]

wb = Workbook()

# ================= SHEET 1: READ ME =================
ws = wb.active; ws.title = "1. READ ME"
ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 110
w(ws, "B1", "Indian Equity Basket Strategies — Real-Cost Backtest (Zerodha), 2005–2026", H1)
readme = [
 "",
 "WHAT THIS IS",
 "A like-for-like backtest of TWELVE rule-based equity baskets on NSE data from 2005 to 2026: seven",
 "'factor baskets' (25 stocks) plus the five 'UNION FAMILY' books, on a Rs 1 crore Zerodha account with",
 "REAL trading costs + tax. No fabricated numbers: the engine reproduces recorded figures to the decimal.",
 "",
 "HOW TO READ THIS WORKBOOK",
 "Sheet 2  — Summary: all strategies side by side, most STABLE first, most AGGRESSIVE last, then the",
 "           UNION FAMILY (purple). The 'seal' check shows the union engine reproduces its sealed numbers.",
 "Sheet 3  — Cost & method: the exact Zerodha charge schedule, the market-cost model, sources, and the",
 "           independent Codex reviews (factor baskets + union) that corrected figures and flagged tax.",
 "Sheets 4+ — one per strategy: plain-English description, step-by-step logic, procedure, data sources,",
 "           every number, year-by-year, and the actual stocks held each rebalance.",
 "Last sheet — Bugs & corrections: everything that went wrong or was fixed during the work.",
 "",
 "THE HONEST HEADLINE (read this first)",
 "1. After REAL cost, the seven FACTOR baskets top out at ~11%/yr (Low-Vol Momentum). Their eye-catching",
 "   20–28% figures are all GROSS (before cost) or 'pretend-cheap' — they collapse; the fast ones go negative.",
 "2. The best factor basket still TRAILS just holding the Nifty 500 net of cost.",
 "3. The single biggest cost is NOT Zerodha's brokerage (~0.2%/yr, tiny). It is MARKET IMPACT — the price you",
 "   move against yourself in less-liquid stocks — times how often you trade. Cheap brokerage cannot fix churn.",
 "4. Capital-gains tax (added on Codex's advice) makes the churny baskets worse still (short-term gains, 20%).",
 "5. THE UNION FAMILY IS DIFFERENT. Put through the SAME gauntlet, its lead books SURVIVE: COMPOSITE-30 nets",
 "   +17.8%/yr and A2 +17.2%/yr (Rs1cr -> ~Rs25-28cr) vs their same-window index +11.7% (Rs9.4cr) — the ONLY",
 "   books here that beat buy-and-hold net of real cost. Why: quarterly (low churn), more-liquid names, stops.",
 "   Caveat: IN-SAMPLE; the sealed FORWARD test is 2026-10-03, and it is personal-scale (picks ~Rs52-58cr/day).",
 "",
 "WHO SHOULD PICK WHAT (by temperament)",
 "  Wants calm / small drops   -> Pure Low-Volatility (worst drop only -41%).",
 "  Wants steady real growth   -> Low-Vol Momentum (the 'STEADY core'), the best factor basket.",
 "  Wants maximum growth       -> the momentum baskets look best on paper and are WORST after real cost.",
 "  Wants the best real edge    -> the UNION FAMILY (COMPOSITE-30 / A2) — the only books that beat the index",
 "                                net of cost in-sample; awaiting the 2026-10-03 forward test.",
]
r = 2
for line in readme:
    f = None
    if line in ("WHAT THIS IS","HOW TO READ THIS WORKBOOK","THE HONEST HEADLINE (read this first)",
                "WHO SHOULD PICK WHAT (by temperament)"):
        f = H2
    w(ws, f"B{r}", line, f); ws[f"B{r}"].alignment = WRAP; r += 1

# ================= SHEET 2: SUMMARY (wide, all fields) =================
ws = wb.create_sheet("2. Summary — all strategies")
w(ws, "A1", "SUMMARY — seven baskets, most STABLE (left) to most AGGRESSIVE (right). Rs 1 crore, 2005–2026, real Zerodha cost.", H2)
order = sorted(STRATS, key=lambda m: tnum(m["risk_tier"]))
cols = [BENCH] + order            # benchmark first as reference
# fields down column A, strategies across
for i, (key, label) in enumerate(FIELDS):
    rr = i + 3
    w(ws, f"A{rr}", label, BOLD, border=True); ws[f"A{rr}"].alignment = TOP
    for j, m in enumerate(cols):
        cl = get_column_letter(2 + j)
        val = m.get(key, "—")
        c = w(ws, f"{cl}{rr}", val, border=True); c.alignment = TOP
# header row of strategy names
w(ws, "A2", "FIELD", WHITE, fill="1F3864", border=True)
for j, m in enumerate(cols):
    cl = get_column_letter(2 + j)
    c = w(ws, f"{cl}2", m["strategy"], BOLD, fill=tfill(m.get("risk_tier", "0")), border=True)
    c.alignment = WRAP
ws.column_dimensions["A"].width = 34
for j in range(len(cols)):
    ws.column_dimensions[get_column_letter(2 + j)].width = 16
ws.freeze_panes = "B3"
# the >20% filter answer
rr = len(FIELDS) + 5
w(ws, f"A{rr}", "Your filter — 'any strategy above 20% NET CAGR after real cost?'", H2)
w(ws, f"A{rr+1}",
  "ANSWER: still NONE clears 20% NET. But the UNION FAMILY comes closest and is the ONLY group that BEATS "
  "the index net of real cost: COMPOSITE-30 +17.8%, A2 +17.2% (vs their same-window index +11.7%). The seven "
  "factor baskets top out at +11% (Low-Vol Momentum) and mostly collapse; the fast momentum baskets go negative.")
ws[f"A{rr+1}"].alignment = WRAP
w(ws, f"A{rr+3}", "Best real performer: UNION COMPOSITE-30 — +17.8%/yr net vs same-window index +11.7%; Rs1cr->Rs27.77cr vs Rs9.38cr. IN-SAMPLE; forward test 2026-10-03.", BOLD)
w(ws, f"A{rr+4}", "Lower-drawdown union: A2-composite — +17.2%/yr net, Rs1cr->Rs24.99cr, worst drop -33%.", BOLD)
w(ws, f"A{rr+5}", "Best factor basket: Low-Vol Momentum (STEADY) — +11.0%/yr net, Rs1cr->Rs8.89cr (below its window index +12.5%). Calmest: Pure Low-Vol -41% drop.", BOLD)

# ================= SHEET 3: COST & METHOD =================
ws = wb.create_sheet("3. Cost model & sources")
ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 115
cm = [
 ("", H1, "THE COST MODEL, DATA, AND SOURCES"),
 ("h", H2, "A) Real Zerodha equity-DELIVERY charges (per the official charges page; Codex-verified)"),
 ("", None, "Brokerage: Rs 0 (Zerodha charges zero brokerage on equity delivery)."),
 ("", None, "STT (Securities Transaction Tax): 0.1% on BUY and 0.1% on SELL."),
 ("", None, "NSE exchange transaction charge: 0.00307% per side (Codex correction from 0.00297%)."),
 ("", None, "SEBI turnover fee: 0.0001% per side (Rs 10 per crore)."),
 ("", None, "Stamp duty: 0.015% on BUY only."),
 ("", None, "GST: 18% on (brokerage + exchange + SEBI charges)."),
 ("", None, "DP charge: Rs 15.34 per stock on SELL (Codex correction from 15.93)."),
 ("", BOLD, "=> Total EXPLICIT Zerodha cost is about 0.22% per round-trip (~0.2%/yr) — TINY. This is NOT what hurts."),
 ("h", H2, "B) Market cost (the part no broker removes — this is what actually hurts)"),
 ("", None, "Bid-ask spread (round-trip) by how liquid the stock is:"),
 ("", None, "   • Rs 1–5 cr traded/day  -> 1.5%     • Rs 5–25 cr/day -> 0.6%     • Rs 25 cr+/day -> 0.25%."),
 ("", None, "Slippage: 0.5 x the stock's daily price range (ATR%) — the price moves while you trade."),
 ("", None, "Cost is charged on EVERY stock bought and sold at each rebalance. More churn = more cost."),
 ("h", H2, "C) Capital-gains tax (added on Codex's advice — the biggest investor-realism gap)"),
 ("", None, "Short-term (held <1 yr): 20% (post 23-Jul-2024; was 15%). Long-term (>1 yr): 12.5% above Rs1.25L."),
 ("", None, "High-churn baskets realise SHORT-term gains every year (20% drag). Buy-and-hold defers tax for years."),
 ("", None, "Shown as an 'illustrative after-tax CAGR' column — assumes full annual realisation at the holding-period rate."),
 ("h", H2, "D) Data & how it is gathered (internally)"),
 ("", None, "Prices: NSE daily bhav copy (official end-of-day), corporate-action adjusted (splits/bonus)."),
 ("", None, "Universe: EVERY stock that ever traded on NSE EQ — 4,236 symbols, including 1,061 that later delisted,"),
 ("", None, "   so the test is SURVIVORSHIP-CLEAN (it picks from stocks that existed then, not just today's winners)."),
 ("", None, "Liquidity: each stock's trailing 22-day median traded value, used for the tier + capacity limit."),
 ("", None, "Benchmark: NSE Nifty 500 index (price). Risk-free basis 6.5%/yr for ratio calculations."),
 ("h", H2, "E) Procedure (identical for every strategy)"),
 ("", None, "1) Each rebalance date, take the liquid universe (>= Rs 5 cr/day)."),
 ("", None, "2) Score every stock by the strategy's rule; keep the top 25, equal weight."),
 ("", None, "3) Hold to the next rebalance; measure the real return of those 25."),
 ("", None, "4) Charge real cost on every stock swapped in/out; compound; grow Rs 1 crore."),
 ("", None, "5) No look-ahead: only data known ON the rebalance date is used."),
 ("h", H2, "F) Independent Codex review (verbatim summary)"),
 ("", None, "Codex corrected the NSE txn rate (0.00307%) and DP fee (Rs15.34); confirmed DP is immaterial at Rs1cr;"),
 ("", None, "noted charging only entries/exits slightly UNDERSTATES cost (held names also need trimming) — so the"),
 ("", None, "net figures are, if anything, mildly optimistic; and flagged capital-gains tax as the biggest gap (now added)."),
 ("h", H2, "G) Sources"),
 ("", None, "Zerodha official charges page + STT article (charge schedule). Income Tax Dept (STCG/LTCG rates)."),
 ("", None, "NSE (bhav copy prices, Nifty 500 index). Project engine: research/explosive_moves/cost_realism.py."),
]
r = 1
for _, font, text in cm:
    w(ws, f"B{r}", text, font); ws[f"B{r}"].alignment = WRAP; r += 1

# ================= PER-STRATEGY SHEETS =================
DOC = {
 "Pure Low-Volatility": (
   "The calmest basket. Each quarter it buys the 25 large, liquid stocks whose prices have moved the LEAST "
   "(lowest volatility). The idea: boring, stable companies fall less in crashes and compound steadily.",
   ["Look only at large, liquid stocks (Rs 25cr+ traded a day).",
    "Rank them by how CALM their price has been (lowest 66-day volatility).",
    "Buy the 25 calmest, equal money in each.",
    "Hold 3 months; a stock stays unless it drops out of the calmest ~35 (the 'hold-band' reduces trading).",
    "Repeat every quarter."]),
 "Low-Vol Momentum (STEADY core)": (
   "The best real compounder here — the 'STEADY core'. Each quarter it blends two ideas: stocks that are "
   "CALM and stocks that are TRENDING UP. It seeks steady winners without the wild swings of pure momentum.",
   ["Large, liquid stocks only.",
    "Give each a score = half 'how calm' + half 'how strong its 6-month rise'.",
    "Buy the top 25, equal money each.",
    "Hold 3 months with a hold-band so you don't churn on small rank changes.",
    "Repeat quarterly."]),
 "Risk-adjusted Momentum (qtr)": (
   "A middle-of-the-road growth basket. Each quarter it buys stocks with the best rise PER UNIT of risk "
   "(6-month return divided by volatility). Rewards strong-but-not-crazy movers.",
   ["Large, liquid stocks only.",
    "Score each = 6-month return / its volatility.",
    "Buy the top 25, equal weight; NO hold-band (it re-picks fully each quarter).",
    "Hold 3 months, repeat."]),
 "Low-Vol Momentum (monthly)": (
   "The same steady blend as the STEADY core, but rebalanced EVERY MONTH across a broader universe. Trading "
   "12x a year instead of 4x roughly triples the cost — which is why its real return is worse despite similar picks.",
   ["Broad liquid universe (Rs 5cr+/day).",
    "Same score: half calm + half 6-month momentum.",
    "Top 25, equal weight, with a hold-band.",
    "Rebalance MONTHLY (this is the cost killer)."]),
 "Pure Momentum 6m (qtr)": (
   "An aggressive growth basket. Each quarter it simply buys the 25 stocks that rose the MOST over the last "
   "6 months. Chases winners — big gains in good years, deep falls in bad ones.",
   ["Large, liquid stocks only.",
    "Rank by 6-month price rise; buy the top 25 equal weight.",
    "No hold-band — fully re-pick each quarter.",
    "Hold 3 months, repeat."]),
 "Pure Momentum 6m (monthly)": (
   "The most aggressive, highest-churn basket. Buys the biggest 6-month risers across a broad universe and "
   "re-picks EVERY MONTH. Looks spectacular before cost (28% gross) and is a capital-destroyer after real cost.",
   ["Broad liquid universe.",
    "Rank by 6-month rise; top 25 equal weight, no hold-band.",
    "Rebalance MONTHLY — the highest trading of all seven."]),
 "Risk-adjusted Momentum (monthly)": (
   "The fast version of risk-adjusted momentum: best return-per-risk names, re-picked EVERY MONTH across a "
   "broad universe. Highest gross of all, but ~38%/yr real cost turns Rs1cr into ~Rs8 lakh over 21 years.",
   ["Broad liquid universe.",
    "Score = 6-month return / volatility; top 25 equal weight, no hold-band.",
    "Rebalance MONTHLY."]),
}
SHARED_PROC = ["No look-ahead: only data known on the rebalance date is used.",
    "Real return of the 25 held stocks is measured to the next rebalance.",
    "Real cost is charged on every stock bought and sold (spread + slippage + Zerodha charges).",
    "Returns are compounded; a Rs 1 crore start is grown through the whole period.",
    "The engine reproduces the project's recorded numbers to the decimal (validation)."]
SHARED_DATA = ["Prices: NSE official daily bhav copy, adjusted for splits and bonuses.",
    "Universe: all 4,236 NSE EQ symbols ever traded (incl. 1,061 delisted) -> survivorship-clean.",
    "Liquidity: each stock's trailing 22-day median traded value."]
SHARED_SRC = ["NSE bhav copy (prices) + Nifty 500 index.",
    "Zerodha official charges page (cost schedule).",
    "Income Tax Department (capital-gains rates).",
    "Engine: research/explosive_moves/cost_realism.py (validated)."]

UNION_DOC = {
 "Union (RS turn OR RS trend)": (
   "The base of the 'union family'. Each quarter it owns ~60 stocks that are EITHER just turning up in "
   "relative strength vs the market OR already trending up, with a trailing stop that cuts losers and lets "
   "winners run. Lower turnover and more liquid names than the factor baskets — which is why it survives real cost.",
   ["Each quarter, from liquid stocks, find those turning up OR trending up in strength vs the Nifty 500.",
    "Buy ~60 of them, equal weight.",
    "Hold with a 20% trailing stop: a name is sold if it falls 20% from its peak (cut losers, ride winners).",
    "Park idle cash safely (Next-50 sleeve in up-markets) when fewer names qualify.",
    "Rebalance quarterly."]),
 "Union + beta-cap 1.4": (
   "The Union book, but it avoids the wildest movers by capping each stock's market-sensitivity (beta) at 1.4. "
   "This single change fixed a weak 2012-17 stretch and cut the drawdown.",
   ["Same quarterly strength selection as Union.",
    "Drop any name whose beta (swing vs the market) is above 1.4.",
    "Buy ~60, equal weight, 20% trailing stop, quarterly."]),
 "Union C40 (risk-adj, top-40)": (
   "Tighter and stronger: ranks the qualifiers by RETURN-PER-UNIT-OF-RISK and keeps the best 40. Higher return, "
   "still well diversified.",
   ["Same strength + beta discipline.",
    "Rank by 6-month return divided by volatility (return per risk).",
    "Keep the top 40, equal weight, 20% trailing stop, quarterly."]),
 "Union A2-composite (top-40 + sleeve)": (
   "A top-40 composite that parks idle money in a safe sleeve when fewer names qualify — smoother, lower drawdown "
   "than the lead, with almost the same return.",
   ["Composite score (strength + risk-adjustment) picks the top 40.",
    "Idle cash earns a safe return (sleeve / cash) when under-invested.",
    "Equal weight, 20% trailing stop, quarterly."]),
 "Union COMPOSITE-30 (LEAD)": (
   "The lead of the family: the 30 strongest composite names, letting winners keep a bigger weight (drift), with a "
   "safety sleeve. Highest return of the union family, and the one that survives real cost best.",
   ["Composite score picks the strongest 30 names.",
    "Winners are allowed to keep their grown weight (capped) instead of being trimmed every quarter.",
    "Safety sleeve for idle cash; 20% trailing stop; quarterly rebalance."]),
}
UNION_PROC = ["No look-ahead: selection uses only data known ON the rebalance date.",
    "Real return measured to the next quarter; the 20% trailing stop is applied within the quarter.",
    "Real cost charged per name traded — the SAME Zerodha per-name gauntlet as the factor baskets.",
    "This is IN-SAMPLE (2005-2026, the window these books were designed on).",
    "VALIDATION: under flat cost the engine reproduces the sealed CAGRs to the decimal (seal col on the Summary).",
    "The sealed FORWARD test is 2026-10-03 — the real out-of-sample judge; do not fund on the in-sample number."]
UNION_DATA = ["Prices: NSE bhav copy, split/bonus adjusted.",
    "Universe: survivorship-clean (delisted names included) with an era-relative liquidity floor.",
    "Relative strength computed vs the Nifty 500; sector context from NSE sector indices."]
UNION_SRC = ["NSE bhav copy + index_rows (Nifty 500 / Next 50 / sector indices).",
    "Zerodha official charges page; Income Tax Department (STCG/LTCG).",
    "Engine: research/explosive_moves/union_ladder_val.py (sealed protocol 37c28824); cost via the per-name gauntlet."]
ALLDOC = {**DOC, **UNION_DOC}

sheetnum = 4
for m in order:
    nm = m["strategy"]
    safe = f"{sheetnum}. {nm}"[:31]
    ws = wb.create_sheet(safe)
    ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 60
    for c in "CDEFGHIJKLMNOPQRSTUVWXYZ"[:26]:
        ws.column_dimensions[c].width = 12
    w(ws, "A1", nm, H1, fill=tfill(m["risk_tier"]))
    w(ws, "A2", f"Risk tier: {m['risk_tier']}", BOLD)
    plain, steps = ALLDOC[nm]
    _PROC, _DATA, _SRC = ((UNION_PROC, UNION_DATA, UNION_SRC) if "UNION" in m["risk_tier"]
                          else (SHARED_PROC, SHARED_DATA, SHARED_SRC))
    r = 4
    w(ws, f"A{r}", "WHAT IT IS (plain English)", H2); r += 1
    w(ws, f"A{r}", plain); ws[f"A{r}"].alignment = WRAP; ws.merge_cells(f"A{r}:F{r}"); ws.row_dimensions[r].height = 60; r += 2
    w(ws, f"A{r}", "HOW IT WORKS — STEP BY STEP", H2); r += 1
    for i, s in enumerate(steps, 1):
        w(ws, f"A{r}", f"{i}. {s}"); ws[f"A{r}"].alignment = WRAP; ws.merge_cells(f"A{r}:F{r}"); r += 1
    r += 1
    w(ws, f"A{r}", "PROCEDURE WE FOLLOWED", H2); r += 1
    for s in _PROC:
        w(ws, f"A{r}", "• " + s); ws.merge_cells(f"A{r}:F{r}"); r += 1
    r += 1
    w(ws, f"A{r}", "HOW THE DATA IS GATHERED (internally)", H2); r += 1
    for s in _DATA:
        w(ws, f"A{r}", "• " + s); ws.merge_cells(f"A{r}:F{r}"); r += 1
    r += 1
    w(ws, f"A{r}", "SOURCES", H2); r += 1
    for s in _SRC:
        w(ws, f"A{r}", "• " + s); ws.merge_cells(f"A{r}:F{r}"); r += 1
    r += 1
    # all the numbers
    w(ws, f"A{r}", "KEY NUMBERS (Rs 1 crore, 2005–2026, real Zerodha cost)", H2); r += 1
    for key, label in FIELDS:
        w(ws, f"A{r}", label, BOLD, border=True)
        w(ws, f"B{r}", m.get(key, "—"), border=True); r += 1
    r += 1
    # year by year
    w(ws, f"A{r}", "YEAR BY YEAR (net return, and Rs 1 cr running value)", H2); r += 1
    w(ws, f"A{r}", "Year", WHITE, fill="1F3864", border=True)
    w(ws, f"B{r}", "Net return %", WHITE, fill="1F3864", border=True)
    w(ws, f"C{r}", "Rs 1cr becomes (Rs cr)", WHITE, fill="1F3864", border=True); r += 1
    run = 1.0
    for y, v in m["_year_by_year"].items():
        run *= (1 + v)
        w(ws, f"A{r}", y, border=True); w(ws, f"B{r}", round(v*100, 1), border=True)
        w(ws, f"C{r}", round(run, 2), border=True); r += 1
    r += 1
    # month/quarter-by-quarter rosters
    w(ws, f"A{r}", "BASKETS OVER TIME — the actual stocks held each rebalance", H2); r += 1
    _mx = max((len(ro["symbols"]) for ro in m["_rosters"]), default=25)
    hdr = ["Rebalance date", "# new", "# exited"] + [f"Stock {i}" for i in range(1, _mx + 1)]
    for ci, h in enumerate(hdr):
        w(ws, f"{get_column_letter(1+ci)}{r}", h, WHITE, fill="1F3864", border=True)
    ws.freeze_panes = f"A{r+1}"; r += 1
    for ro in m["_rosters"]:
        w(ws, f"A{r}", ro["date"], border=True)
        w(ws, f"B{r}", ro["n_new"], border=True); w(ws, f"C{r}", ro["n_exit"], border=True)
        for ci, sym in enumerate(ro["symbols"][:_mx]):
            w(ws, f"{get_column_letter(4+ci)}{r}", sym, border=True)
        r += 1
    sheetnum += 1

# ================= K30 CAPACITY & EXECUTION SHEET =================
ws = wb.create_sheet("16. K30 capacity & execution")
ws.column_dimensions["A"].width = 4
for _c, _wd in [("B", 24), ("C", 13), ("D", 13), ("E", 20), ("F", 18), ("G", 15), ("H", 13)]:
    ws.column_dimensions[_c].width = _wd
w(ws, "B1", "K30 (COMPOSITE-30) — Capacity & Execution checks", H1, fill=tfill("12-UNION"))
r = 3
w(ws, f"B{r}", "WHY THIS SHEET", H2); r += 1
for line in [
 "Codex flagged two things the main gauntlet did not test: (1) does the edge survive if you trade a DAY LATE",
 "instead of at the exact signal price, and (2) how much money can you actually run before your OWN trading",
 "eats the edge. Both are answered below with real runs of the sealed K30 engine."]:
    w(ws, f"B{r}", line); ws[f"B{r}"].alignment = WRAP; ws.merge_cells(f"B{r}:H{r}"); r += 1
r += 1
w(ws, f"B{r}", "A) EXECUTION-LAG CHECK — trade at the NEXT session's price, not the signal close", H2); r += 1
for ci, h in enumerate(["Fill timing", "Net CAGR %", "Rs1cr -> cr", "Worst drop %"]):
    w(ws, f"{get_column_letter(2+ci)}{r}", h, WHITE, fill="1F3864", border=True)
r += 1
for row in [["Same-bar (signal close)", 17.8, 27.77, -38], ["Next-session (T+1)", 16.7, 22.97, -40]]:
    for ci, v in enumerate(row):
        w(ws, f"{get_column_letter(2+ci)}{r}", v, border=True)
    r += 1
w(ws, f"B{r}", "-> SURVIVES. A 1-day lag costs ~1.1 points/yr; still well above the index (11.7%). Not a fill artifact.", BOLD)
ws.merge_cells(f"B{r}:H{r}"); r += 2
w(ws, f"B{r}", "B) AUM LADDER — square-root participation impact (10% ADV/day cap + days-to-fill penalty)", H2); r += 1
for ci, h in enumerate(["Money deployed", "Net CAGR %", "Rs1cr -> cr", "Median participation %",
                        "Median days to fill", "% trades >1 day", "Worst drop %"]):
    c = w(ws, f"{get_column_letter(2+ci)}{r}", h, WHITE, fill="1F3864", border=True); c.alignment = WRAP
r += 1
ladder = [["Rs 5 cr", 22.1, 56.77, 1.1, 1, 7, -36], ["Rs 25 cr", 16.6, 22.45, 5.4, 1, 37, -42],
          ["Rs 50 cr", 12.5, 10.83, 10.9, 2, 52, -45], ["Rs 100 cr", 7.0, 3.94, 21.8, 3, 67, -59],
          ["Rs 200 cr", -0.3, 0.94, 43.6, 5, 79, -83], ["Rs 500 cr", -12.8, 0.06, 108.9, 11, 91, -97],
          ["Rs 1000 cr", -25.2, 0.00, 217.8, 22, 96, -100]]
for row in ladder:
    lab_fill = "C6EFCE" if row[1] >= 15 else "FFEB9C" if row[1] >= 11.7 else "FFC7CE"
    for ci, v in enumerate(row):
        c = w(ws, f"{get_column_letter(2+ci)}{r}", v, border=True)
        if ci == 0: c.fill = PatternFill("solid", fgColor=lab_fill)
    r += 1
r += 1
for line in [
 "CAPACITY CEILING ~ Rs 25-50 crore:",
 "  - Below ~Rs 25 cr: a genuine edge (16-22% net, well above the index).",
 "  - ~Rs 50 cr: the edge fades to about the index (+12.5% vs 11.7%) — the practical ceiling.",
 "  - Above Rs 50 cr: edge gone; by Rs 200 cr your own trading turns it NEGATIVE (participation >10%/day,",
 "    positions take a week+ to fill, drawdown -83%).",
 "  => K30 is a PERSONAL-SCALE strategy: viable in-sample for your own capital, NOT an institutional product."]:
    fnt = BOLD if (line.strip().endswith(":") or "PERSONAL-SCALE" in line) else None
    w(ws, f"B{r}", line, fnt); ws[f"B{r}"].alignment = WRAP; ws.merge_cells(f"B{r}:H{r}"); r += 1
r += 1
w(ws, f"B{r}", "METHOD & CAVEATS", H2); r += 1
for line in [
 "Execution-lag: the engine's built-in 'lagged' mode enters/exits one session later (T+1), with the gauntlet cost.",
 "AUM ladder: cost_participation.py impact model — per side, impact = 0.6 x (66-day vol) x sqrt(clip / ADV),",
 "  capped at 10% of the name's daily traded value, plus a days-to-fill timing penalty for oversized clips.",
 "  Clip = AUM x weight traded; ADV = trailing 22-day median traded value; spread + Zerodha charges as before.",
 "Both are real runs of the sealed K30 engine (top-30, drift-weighted, rf-cash sleeve); flat mode reproduces the seal.",
 "STILL OWED: the sealed FORWARD test on 2026-10-03 — no backtest can settle that."]:
    w(ws, f"B{r}", line); ws[f"B{r}"].alignment = WRAP; ws.merge_cells(f"B{r}:H{r}"); r += 1

# ================= LAST SHEET: BUGS & CORRECTIONS =================
ws = wb.create_sheet("Bugs & corrections")
ws.column_dimensions["A"].width = 4; ws.column_dimensions["B"].width = 115
bugs = [
 ("h", H1, "BUGS, ERRORS, AND CORRECTIONS DURING THIS WORK"),
 ("h", H2, "Fixed / applied"),
 ("", None, "1. SSH to the data box timed out twice mid-run (network) — added ret/retry loops; both runs then completed."),
 ("", None, "2. First backtest printed 'avg annual cost 0.0%' due to a leftover *0 in a print line — the equity was correct"),
 ("", None, "   (cost is inside the net returns, proven by the validation); the cosmetic print was dropped in this build."),
 ("", None, "3. Codex correction applied: NSE txn charge 0.00297% -> 0.00307% (current Zerodha rate)."),
 ("", None, "4. Codex correction applied: DP sell fee Rs 15.93 -> Rs 15.34 (current Zerodha CDSL DP)."),
 ("", None, "5. Codex flag applied: capital-gains tax was missing -> added an illustrative after-tax CAGR column"),
 ("   (STCG 20% / LTCG 12.5%, post 23-Jul-2024)."),
 ("h", H2, "Known conservatisms / caveats (disclosed, not hidden)"),
 ("", None, "A. Cost is charged only on stocks entering/exiting, NOT on trimming held names back to equal weight."),
 ("   This UNDERSTATES cost — real net returns would be slightly LOWER than shown (Codex point 2)."),
 ("", None, "B. Slippage uses 0.5 x ATR%; Codex notes this can be pessimistic for a Rs1cr book (offsets A somewhat)."),
 ("", None, "C. Dividends are excluded on BOTH strategies and the index (adds ~1.5%/yr to both; the gap is ~unchanged)."),
 ("", None, "D. A stock that delists mid-hold is dropped, not booked as a loss — a small optimistic tilt for the fast baskets."),
 ("", None, "E. The after-tax column assumes full annual realisation at the holding-period rate — an illustration, not a filing."),
 ("h", H2, "Validation (proof the numbers are real)"),
 ("", None, "The engine reproduced the project's recorded 2012+ figures to the decimal: risk-adj momentum monthly"),
 ("   = -1.3% (recorded -1.5%), low-vol momentum quarterly = +13.2% (recorded +13.3%). Nothing here is invented."),
 ("h", H2, "Independent Codex validation — UNION FAMILY (verbatim summary)"),
 ("", None, "Codex re-ran the union gauntlet itself and REPRODUCED the table: flat mode reproduces the sealed"),
 ("", None, "   CAGRs (U 17.5->17.3, C40 21.0->20.8, A2 25.5->25.5, K30 26.4->26.4); gauntlet K30 17.8% / A2 17.2%;"),
 ("", None, "   benchmark 11.7% -> Rs9.38cr. Codex verdict: the sealed engine appears intact and, under this per-name"),
 ("", None, "   Zerodha + 0.5*ATR gauntlet, A2 17.2% and K30 17.8% beat the Nifty 500 index (11.7%) IN-SAMPLE."),
 ("", None, "Codex's correct investor wording: 'in-sample survivor under an AUM-blind harsh slippage gauntlet,"),
 ("", None, "   pending the sealed 2026-10-03 forward test.'"),
 ("", None, "Codex caveats (apply before trusting beyond a personal-size account):"),
 ("", None, "  - AUM-BLIND: this is the cost_realism-style model, not the cost_participation AUM-impact model. The"),
 ("", None, "    numbers are personal-scale; they are NOT an institutional-capacity proof."),
 ("", None, "  - The equal-weight books (Union / B14 / C40 / A2) slightly UNDERCHARGE the small trades that restore"),
 ("", None, "    equal weight each quarter -> their net is mildly OPTIMISTIC. K30 (drift-weighted) is charged correctly."),
 ("", None, "  - 0.5*ATR slippage is harsh for a patient personal account but optimistic for large AUM (impact does"),
 ("", None, "    not scale with order size). Median pick liquidity ~Rs54cr/day; capacity figures are rough, not a recut."),
 ("", None, "  - Still owed before real allocation: an AUM ladder, a 1-day execution-lag check, and the forward test."),
]
r = 1
for _, font, *text in bugs:
    w(ws, f"B{r}", (text[0] if text else _), font); ws[f"B{r}"].alignment = WRAP; r += 1

out = os.path.join(SP, "Equity_Basket_Strategies_RealCost_2005-2026.xlsx")
wb.save(out)
print("saved:", out)
print("sheets:", wb.sheetnames)

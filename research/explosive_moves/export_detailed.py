"""Detailed trade log: rank-based ENTRY + explicit trailing STOP-LOSS exit.

Entry: monthly, the top-25 by risk-adjusted momentum (6m return / volatility), >=5cr.
Per position we log WHY picked (rank, score, 6m return, volatility, 5y extension) and
manage an explicit stop-loss on DAILY prices:
  - initial SL = -18% from buy
  - raise SL to breakeven once +15% in profit
  - then trail SL 20% under the highest price once +30% in profit (ratchets up only)
Exit = whichever comes first: SL hit, or the stock drops out of the top-25 at a monthly
re-rank. Logs initial SL, how it moved, peak, exit trigger, exit basis, return.
"""
import csv
import numpy as np
from pathlib import Path
from explosive_moves.embase import load_symbol_cache
from explosive_moves.strategies import CR
from explosive_moves.metrics import index_series
from explosive_moves.factory import build_tables, sig_riskadj, eqstats, COST_PS

OUT = Path("/opt/hermes/research/explosive_moves/out"); OUT.mkdir(exist_ok=True)
TOPN, SL0, BE_AT, TRAIL_AT, TRAIL = 25, 0.18, 0.15, 0.30, 0.20
RT = 2 * COST_PS

cache = load_symbol_cache()
D2I = {s: {d: i for i, d in enumerate(A["date"])} for s, A in cache.items()}
dts, cl = index_series("Nifty 50")
cal = [d for d in dts if d >= "2012-06-01"]
rebal_idx = list(range(0, len(cal), 22))
tables = build_tables(cache, cal, rebal_idx, 5 * CR)

# 1) per-rebalance top-25 with entry metrics + ranks
picks_by_r = []
for t in tables:
    score, _ = sig_riskadj(t)
    valid = ~np.isnan(score) & ~np.isnan(t["fwd"])
    idx = np.where(valid)[0]
    order = idx[np.argsort(score[idx])[::-1]]
    top = order[:TOPN]
    n = len(idx)
    rankmap = {t["syms"][j]: (r + 1, float(score[j]), float(t["mom6"][j]),
                              float(t["vol"][j]), float(t["ext5y"][j])) for r, j in enumerate(order)}
    picks_by_r.append({"d0": t["d0"], "top": set(t["syms"][j] for j in top), "info": rankmap})

# 2) build stints (entry rebalance -> first rebalance no longer in top-25)
stints = []
open_pos = {}
for r, P in enumerate(picks_by_r):
    if P["d0"] < "2019-01-01":
        # still seed holdings so a stock entering exactly at boundary is fine
        pass
    for sym in P["top"]:
        if sym not in open_pos and P["d0"] >= "2019-01-01":
            open_pos[sym] = r
    for sym in list(open_pos):
        if sym not in P["top"]:
            stints.append((sym, open_pos[sym], r)); del open_pos[sym]
last_r = len(picks_by_r) - 1
for sym, r0 in open_pos.items():
    stints.append((sym, r0, None))   # still held

# 3) simulate SL per stint on daily data
rows = []
daily_ret = {}     # date -> list of position daily returns (for approx equity)
for sym, r0, r1 in stints:
    A = cache[sym]; d2i = D2I[sym]
    ac, ah, al = A["adj_close"], A["adj_high"], A["adj_low"]
    ed = picks_by_r[r0]["d0"]; i0 = d2i.get(ed)
    if i0 is None:
        continue
    entry = float(ac[i0])
    rank, score, mom6, vol, ext5y = picks_by_r[r0]["info"][sym]
    drop_date = picks_by_r[r1]["d0"] if r1 is not None else picks_by_r[last_r]["d0"]
    i_drop = d2i.get(drop_date, min(i0 + 22, len(ac) - 1))
    sl = entry * (1 - SL0); peak = entry; be = trailed = False
    exit_i, trig = i_drop, ("RANK_DROP" if r1 is not None else "OPEN")
    prev = entry
    for t in range(i0 + 1, i_drop + 1):
        if al[t] <= sl:                         # stop hit
            exit_i, trig = t, "STOP_LOSS"; break
        daily_ret.setdefault(A["date"][t], []).append(ac[t] / prev - 1); prev = ac[t]
        peak = max(peak, ah[t]); g = peak / entry - 1
        if g >= BE_AT:
            sl = max(sl, entry); be = True
        if g >= TRAIL_AT:
            sl = max(sl, peak * (1 - TRAIL)); trailed = True
    exit_px = min(float(A["adj_open"][exit_i]), sl) if trig == "STOP_LOSS" else float(ac[exit_i])
    ret = exit_px / entry - 1 - RT
    peak_g = peak / entry - 1
    sl_path = f"-18%{' ->BE@+15%' if be else ''}{' ->20%-trail' if trailed else ''}"
    why_in = (f"Rank #{rank} of {len(picks_by_r[r0]['info'])} by 6m-return/volatility | "
              f"6m {mom6*100:+.0f}%, vol ~{vol*np.sqrt(252)*100:.0f}%/yr" +
              (f", 5y {ext5y*100:+.0f}% (not over-extended)" if ext5y == ext5y and ext5y < 2
               else (f", 5y {ext5y*100:+.0f}%" if ext5y == ext5y else "")))
    if trig == "STOP_LOSS":
        why_out = f"Stop-loss hit at {exit_px:.0f}; had peaked +{peak_g*100:.0f}%, SL trailed to {sl:.0f}"
    elif trig == "RANK_DROP":
        why_out = f"Dropped out of top-25 (momentum faded); sold at close {exit_px:.0f}"
    else:
        why_out = "Still held (in top-25 currently)"
    rows.append([sym, A["date"][i0], round(entry, 1), rank, round(score, 2),
                 round(mom6 * 100, 0), round(vol * np.sqrt(252) * 100, 0),
                 (round(ext5y * 100, 0) if ext5y == ext5y else "n/a"), why_in,
                 round(entry * (1 - SL0), 1), sl_path, round(peak, 1), round(sl, 1),
                 A["date"][exit_i], round(exit_px, 1), trig, why_out, round(ret * 100, 1)])

# 4) approx equity of this SL version
dates_sorted = sorted(daily_ret)
pr = np.array([np.mean(daily_ret[d]) for d in dates_sorted])   # equal-weight among alive
eq = np.cumprod(1 + pr); peak = np.maximum.accumulate(eq); dd = (eq / peak - 1).min()
yrs = len(eq) / 252; cagr = eq[-1] ** (1 / yrs) - 1
wins = sum(1 for r in rows if r[-1] > 0)
print(f"SL-overlay version: ~CAGR {cagr*100:.1f}%  MaxDD {dd*100:.1f}%  "
      f"({len(rows)} calls, {wins/len(rows)*100:.0f}% win)  [pure rank version was 34.7% / -30%]")
sl_n = sum(1 for r in rows if r[15] == "STOP_LOSS"); rk = sum(1 for r in rows if r[15] == "RANK_DROP")
print(f"Exits: {sl_n} by stop-loss, {rk} by rank-drop, {len(rows)-sl_n-rk} still open")

HDR = ["Stock", "Buy date", "Buy price", "Entry rank", "Risk-adj score", "6m return % at entry",
       "Volatility %/yr at entry", "5y return %", "WHY PICKED", "Initial SL", "SL path",
       "Peak price", "Final SL", "Sell date", "Sell price", "Exit trigger", "WHY EXITED", "Return %"]
rows.sort(key=lambda r: r[1])   # by buy date
with open(OUT / "riskadj_detailed.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(HDR); w.writerows(rows)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = Workbook(); ws = wb.active; ws.title = "Detailed Calls"
ws.append(HDR)
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="305496")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
for r in rows:
    ws.append(r)
ws.freeze_panes = "A2"
widths = [11, 11, 9, 6, 8, 9, 10, 8, 46, 9, 22, 10, 9, 11, 9, 11, 50, 8]
for i, wd in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else "A" + chr(64 + i - 26)].width = wd
wb.save(OUT / "RiskAdjMomentum_detailed_2019.xlsx")
print("Wrote out/RiskAdjMomentum_detailed_2019.xlsx +", len(rows), "rows")
